"""
report_generator.py
===================
Generates a standardized performance analysis report that the user can download.

The report includes:
  - Thesis recap and execution parameters
  - Per-model strategy summary and holdings
  - Performance metrics (total return, annualized, Sharpe-like, max drawdown)
  - Head-to-head comparison table across all models
  - Transaction logs per model
  - Benchmark comparison data

Output formats:
  - Markdown (.md) for readable report
  - Excel (.xlsx) for data-heavy analysis
"""

from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import numpy as np


# ── Styling constants for Excel ────────────────────────────────────────────────
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill("solid", fgColor="2F5496")
MONEY_FMT = '$#,##0.00'
PCT_FMT = '0.00%'
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_header(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


# ── Performance metrics calculation ────────────────────────────────────────────

def compute_metrics(portfolio: dict, prices: dict, benchmark_series: pd.Series = None,
                    risk_free_rate: float = 0.05) -> dict:
    """
    Compute key performance metrics for a single portfolio.

    Args:
        portfolio:       Portfolio state dict from transaction_store.
        prices:          Current {ticker: price} for valuation.
        benchmark_series: Pandas Series of benchmark daily closes (optional).
        risk_free_rate:  Annualized risk-free rate for Sharpe calculation.

    Returns:
        Dict of computed metrics.
    """
    initial = portfolio["initial_capital"]
    # Current value
    current_value = portfolio["cash"]
    for ticker, info in portfolio["holdings"].items():
        p = prices.get(ticker.upper())
        if p is not None:  # Treat $0 as valid; only skip truly missing
            current_value += info["shares"] * p
    current_value = round(current_value, 2)

    total_return = (current_value - initial) / initial if initial else 0
    total_return_dollar = current_value - initial

    # Time elapsed for annualization
    start_str = portfolio.get("start_date", datetime.now().strftime("%Y-%m-%d"))
    try:
        start_dt = datetime.strptime(start_str, "%Y-%m-%d")
    except ValueError:
        start_dt = datetime.now()
    days_elapsed = max((datetime.now() - start_dt).days, 1)
    years_elapsed = days_elapsed / 365.25

    # Annualized return (guard against zero/negative to avoid math errors)
    if years_elapsed > 0 and current_value > 0 and initial > 0:
        ratio = current_value / initial
        # Fractional exponent of a negative number is undefined — fall back
        if ratio > 0:
            annualized = ratio ** (1 / years_elapsed) - 1
        else:
            annualized = total_return
    else:
        annualized = total_return

    # Max drawdown from snapshots
    snapshots = portfolio.get("performance_snapshots", [])
    max_drawdown = 0.0
    if snapshots:
        values = [s["portfolio_value"] for s in snapshots]
        peak = values[0]
        for v in values:
            if v > peak:
                peak = v
            dd = (peak - v) / peak if peak > 0 else 0
            if dd > max_drawdown:
                max_drawdown = dd

    # Simple volatility estimate from snapshots
    volatility = 0.0
    # Pre-compute clean daily returns (reused by volatility, Sortino, win rate, etc.)
    clean_daily_returns = pd.Series(dtype=float)
    if len(snapshots) > 2:
        vals = pd.Series([float(s["portfolio_value"]) for s in snapshots])
        raw_returns = vals.pct_change().dropna()
        clean_daily_returns = raw_returns.replace([np.inf, -np.inf], np.nan).dropna()
        if len(clean_daily_returns) > 0:
            volatility = float(clean_daily_returns.std() * np.sqrt(252))

    # Sharpe-like ratio (using snapshot-based vol)
    sharpe = (annualized - risk_free_rate) / volatility if volatility > 0 else 0.0

    # Sortino ratio — penalises only downside volatility
    sortino = 0.0
    downside_dev = 0.0
    try:
        if len(clean_daily_returns) > 1:
            daily_rf = (1 + risk_free_rate) ** (1 / 252) - 1
            negative_excess = clean_daily_returns[clean_daily_returns < daily_rf] - daily_rf
            if len(negative_excess) > 0:
                dsd = float(np.sqrt((negative_excess ** 2).mean()) * np.sqrt(252))
                if dsd > 0 and np.isfinite(dsd):
                    downside_dev = dsd
                    s = (annualized - risk_free_rate) / dsd
                    if np.isfinite(s):
                        sortino = s
    except Exception:
        sortino = 0.0
        downside_dev = 0.0

    # Calmar ratio — annualized return / max drawdown
    calmar = annualized / max_drawdown if max_drawdown > 0 else 0.0

    # Win rate — percentage of days with positive returns
    win_rate = 0.0
    profit_factor = 0.0
    if len(clean_daily_returns) > 0:
        wins = clean_daily_returns[clean_daily_returns > 0]
        losses = clean_daily_returns[clean_daily_returns < 0]
        win_rate = len(wins) / len(clean_daily_returns)
        # Profit factor — sum of gains / abs(sum of losses)
        loss_sum = abs(float(losses.sum())) if len(losses) > 0 else 0.0
        if loss_sum > 0:
            profit_factor = float(wins.sum()) / loss_sum

    # Number of trades
    num_trades = len(portfolio.get("transactions", []))
    num_holdings = len(portfolio.get("holdings", {}))

    # Best and worst holdings
    holding_returns = {}
    for ticker, info in portfolio["holdings"].items():
        p = prices.get(ticker.upper())
        if p is not None and info["avg_cost"] > 0:
            holding_returns[ticker] = (p - info["avg_cost"]) / info["avg_cost"]

    best_holding = max(holding_returns, key=holding_returns.get) if holding_returns else "N/A"
    worst_holding = min(holding_returns, key=holding_returns.get) if holding_returns else "N/A"
    best_return = holding_returns.get(best_holding, 0)
    worst_return = holding_returns.get(worst_holding, 0)

    # AI generation cost
    ai_usage = portfolio.get("ai_usage") or {}
    ai_cost = ai_usage.get("estimated_cost_usd", 0.0)
    ai_tokens = ai_usage.get("total_tokens", 0)
    net_return_dollar = total_return_dollar - ai_cost
    net_return_pct = net_return_dollar / initial if initial else 0

    return {
        "current_value": current_value,
        "total_return_pct": total_return,
        "total_return_dollar": total_return_dollar,
        "annualized_return": annualized,
        "max_drawdown": max_drawdown,
        "volatility": volatility,
        "sharpe_ratio": sharpe,
        "sortino_ratio": sortino,
        "calmar_ratio": calmar,
        "downside_deviation": downside_dev,
        "days_elapsed": days_elapsed,
        "num_trades": num_trades,
        "num_holdings": num_holdings,
        "best_holding": best_holding,
        "best_holding_return": best_return,
        "worst_holding": worst_holding,
        "worst_holding_return": worst_return,
        "cash_remaining": portfolio["cash"],
        "invested_pct": 1 - (portfolio["cash"] / current_value) if current_value > 0 else 0,
        # AI cost fields
        "ai_cost_usd": ai_cost,
        "ai_tokens": ai_tokens,
        "net_return_dollar": net_return_dollar,
        "net_return_pct": net_return_pct,
        # Win/loss fields
        "win_rate": win_rate,
        "profit_factor": profit_factor,
    }


# ── Comparative narrative builder ──────────────────────────────────────────────

# Well-known ETF tickers for classification heuristic
_KNOWN_ETFS = {
    "SPY", "QQQ", "IWM", "VTI", "VOO", "VEA", "VWO", "BND", "AGG",
    "XLK", "XLF", "XLE", "XLV", "XLI", "XLP", "XLRE", "XLY", "XLU",
    "XLB", "XLC", "ARKK", "ARKW", "ARKG", "SOXX", "SMH", "TAN",
    "ICLN", "IBIT", "GLD", "SLV", "TLT", "HYG", "LQD", "DIA",
    "IVV", "SCHD", "VIG", "VXUS", "EEM", "EFA", "SOXL", "TQQQ",
}


def _classify_etf(alloc: dict) -> bool:
    """Heuristic: is this allocation an ETF?"""
    etf_keywords = {"ETF", "FUND", "INDEX", "TRUST"}
    name = alloc.get("name", "").upper()
    return (
        any(kw in name for kw in etf_keywords)
        or alloc["ticker"].upper() in _KNOWN_ETFS
    )


def _build_model_profile(port: dict) -> dict:
    """
    Extract a structured profile of how one model converted the thesis,
    suitable for comparison and narrative generation.
    """
    sd = port.get("strategy_data", {})
    allocs = sd.get("allocations", [])
    sorted_allocs = sorted(allocs, key=lambda a: a.get("weight", 0), reverse=True)

    etf_count = sum(1 for a in allocs if _classify_etf(a))
    stock_count = len(allocs) - etf_count
    top_3 = sorted_allocs[:3]
    total_weight = sum(a.get("weight", 0) for a in allocs)

    # Concentration: weight of the single largest position
    top_weight = sorted_allocs[0]["weight"] if sorted_allocs else 0
    # Overlap tickers (used for cross-model comparison)
    tickers = {a["ticker"].upper() for a in allocs}

    return {
        "label": port.get("model_label", "Unknown"),
        "strategy_name": port.get("strategy_name", "N/A"),
        "risk_level": sd.get("risk_level", "Unknown"),
        "time_horizon": sd.get("time_horizon", "N/A"),
        "rationale": sd.get("rationale", ""),
        "rebalancing": sd.get("rebalancing_notes", "N/A"),
        "num_positions": len(allocs),
        "etf_count": etf_count,
        "stock_count": stock_count,
        "top_3": top_3,
        "top_weight": top_weight,
        "total_weight_allocated": total_weight,
        "cash_pct": max(0, 1 - total_weight),
        "tickers": tickers,
        "allocs": sorted_allocs,
    }


def _build_comparative_narrative(portfolios: dict) -> str:
    """
    Generate a multi-paragraph comparative summary that highlights how
    the models diverged in converting the same thesis into trading strategies.

    Covers:
      - Risk posture differences
      - Instrument selection (ETFs vs stocks, concentration vs diversification)
      - Ticker overlap / divergence
      - Time horizon alignment
      - Overall interpretation style

    Returns:
        A Markdown-formatted narrative string.
    """
    profiles = {label: _build_model_profile(port) for label, port in portfolios.items()}

    if len(profiles) < 2:
        # Only one model — no comparison possible; give a single-model summary
        p = list(profiles.values())[0]
        return (
            f"**{p['label']}** interpreted the thesis as a **{p['risk_level'].lower()}** "
            f"opportunity with a **{p['time_horizon'].lower()}** time horizon. "
            f"It constructed a {p['num_positions']}-position portfolio "
            f"({p['etf_count']} ETFs, {p['stock_count']} individual stocks) "
            f"with the largest position at {p['top_weight']*100:.0f}% of capital. "
            f"{p['rationale']}"
        )

    paras = []

    # ── Paragraph 1: Risk posture comparison ──────────────────────────────
    risk_levels = {l: p["risk_level"] for l, p in profiles.items()}
    unique_risks = set(risk_levels.values())
    if len(unique_risks) == 1:
        paras.append(
            f"All models converged on a **{list(unique_risks)[0].lower()}** risk posture, "
            f"suggesting the thesis naturally lends itself to this risk tier."
        )
    else:
        risk_parts = [f"**{l}** chose a **{r.lower()}** approach" for l, r in risk_levels.items()]
        paras.append(
            "The models diverged significantly in risk posture: "
            + "; ".join(risk_parts) + ". "
            "This divergence highlights how the same narrative can be interpreted as either "
            "a defensive or an opportunistic thesis depending on the model's reasoning style."
        )

    # ── Paragraph 2: Instrument mix & concentration ───────────────────────
    mix_parts = []
    for label, p in profiles.items():
        concentration = "concentrated" if p["top_weight"] >= 0.30 else "diversified"
        vehicle = (
            "ETF-heavy" if p["etf_count"] > p["stock_count"]
            else "stock-heavy" if p["stock_count"] > p["etf_count"]
            else "balanced between ETFs and stocks"
        )
        mix_parts.append(
            f"**{label}** built a {concentration}, {vehicle} portfolio "
            f"with {p['num_positions']} positions"
        )
    paras.append(
        "In terms of construction style: "
        + "; ".join(mix_parts) + "."
    )

    # ── Paragraph 3: Ticker overlap / divergence ──────────────────────────
    all_ticker_sets = [p["tickers"] for p in profiles.values()]
    common = set.intersection(*all_ticker_sets) if all_ticker_sets else set()
    all_unique = set.union(*all_ticker_sets) if all_ticker_sets else set()

    if common:
        overlap_pct = len(common) / len(all_unique) * 100 if all_unique else 0
        paras.append(
            f"The models share **{len(common)}** ticker(s) in common "
            f"({', '.join(sorted(common))}), representing "
            f"**{overlap_pct:.0f}%** of all unique positions across portfolios. "
            + (
                "This substantial overlap suggests strong consensus on these names."
                if overlap_pct > 40
                else "The limited overlap indicates each model found a distinct path "
                     "to express the same thesis."
            )
        )
    else:
        paras.append(
            "Interestingly, the models chose **entirely different tickers** — there is "
            "zero overlap between portfolios. This suggests the thesis is broad enough "
            "that multiple valid implementations exist with no consensus picks."
        )

    # ── Paragraph 4: Time horizon & rebalancing ──────────────────────────
    horizons = {l: p["time_horizon"] for l, p in profiles.items()}
    unique_horizons = set(horizons.values())
    if len(unique_horizons) == 1:
        paras.append(
            f"All models agree on a **{list(unique_horizons)[0].lower()}** time horizon."
        )
    else:
        hz_parts = [f"{l} targets {h.lower()}" for l, h in horizons.items()]
        paras.append(
            "Time horizons differ: " + "; ".join(hz_parts) + ". "
            "This means the strategies are optimizing for different holding periods, "
            "which directly affects how performance should be evaluated."
        )

    return "\n\n".join(paras)


# ── Markdown report generation ─────────────────────────────────────────────────

def generate_markdown_report(
    thesis: str,
    portfolios: dict,
    all_prices: dict,
    benchmark_ticker: str,
    initial_capital: float,
    start_date: str,
) -> str:
    """
    Generate a comprehensive Markdown performance report.

    Args:
        thesis:          The original investment thesis text.
        portfolios:      Dict of {model_label: portfolio_state_dict}.
        all_prices:      Combined {ticker: price} for all models.
        benchmark_ticker: Benchmark symbol used.
        initial_capital: Starting capital.
        start_date:      Backtest / execution start date.

    Returns:
        Markdown string of the full report.
    """
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    lines = []

    lines.append("# Investment Strategy Performance Report")
    lines.append(f"\n**Generated:** {now}")
    lines.append(f"**Initial Capital:** ${initial_capital:,.2f}")
    lines.append(f"**Start Date:** {start_date}")
    lines.append(f"**Benchmark:** {benchmark_ticker}")
    lines.append(f"**Models Compared:** {len(portfolios)}")

    lines.append("\n---\n")
    lines.append("## Investment Thesis")
    lines.append(f"\n{thesis}")

    # ── NEW: Thesis-to-Strategy Interpretation per model ──────────────────
    lines.append("\n---\n")
    lines.append("## How Each Model Interpreted the Thesis")
    lines.append("")
    lines.append(
        "Each model independently received the same thesis and translated it "
        "into an actionable trading strategy. Below is a comparative summary "
        "followed by a detailed breakdown of each model's approach."
    )

    # ── Comparative narrative: synthesizes key differences across models ──
    lines.append("\n### Comparative Analysis\n")
    lines.append(_build_comparative_narrative(portfolios))

    # ── Per-model detailed breakdowns ─────────────────────────────────────
    for label, port in portfolios.items():
        sd = port.get("strategy_data", {})
        allocs = sd.get("allocations", [])
        profile = _build_model_profile(port)

        # ── Model header & high-level interpretation ──
        lines.append(f"\n### 🤖 {label} → *{port.get('strategy_name', 'N/A')}*\n")

        # Short narrative: HOW this model converted the thesis into action
        concentration = "concentrated" if profile["top_weight"] >= 0.30 else "well-diversified"
        vehicle_desc = (
            "primarily through ETFs" if profile["etf_count"] > profile["stock_count"]
            else "primarily through individual stocks" if profile["stock_count"] > profile["etf_count"]
            else "through a balanced mix of ETFs and individual stocks"
        )
        top_pick = profile["top_3"][0] if profile["top_3"] else None
        top_desc = (
            f", anchored by a {top_pick['weight']*100:.0f}% position in "
            f"{top_pick['ticker']} ({top_pick.get('name', top_pick['ticker'])})"
            if top_pick else ""
        )

        lines.append(
            f"**Thesis → Strategy Translation:** {label} interpreted the thesis as a "
            f"**{profile['risk_level'].lower()}** opportunity and expressed it "
            f"{vehicle_desc} over a **{profile['time_horizon'].lower()}** horizon. "
            f"It built a {concentration} portfolio of {profile['num_positions']} positions"
            f"{top_desc}. "
            f"{profile['rationale']}"
        )
        lines.append("")

        lines.append(f"**Risk Posture:** {profile['risk_level']} | "
                     f"**Time Horizon:** {profile['time_horizon']}")
        lines.append("")

        # Allocation summary
        sorted_allocs = profile["allocs"]
        top_picks = sorted_allocs[:3]
        top_picks_str = ", ".join(
            f"**{a['ticker']}** ({a.get('name', a['ticker'])}, {a['weight']*100:.0f}%)"
            for a in top_picks
        )
        remaining = len(sorted_allocs) - 3
        if remaining > 0:
            top_picks_str += f", and {remaining} other position(s)"

        lines.append(
            f"**Allocation Approach:** {profile['num_positions']} positions "
            f"({profile['etf_count']} ETF(s), {profile['stock_count']} individual stock(s)). "
            f"Top holdings: {top_picks_str}."
        )
        if profile["cash_pct"] > 0.01:
            lines.append(
                f" Approximately {profile['cash_pct']*100:.0f}% left as cash."
            )
        lines.append("")

        # Per-pick reasoning
        lines.append("**Position-Level Reasoning:**\n")
        for a in sorted_allocs:
            pick_rationale = a.get("rationale", "No specific reason given.")
            lines.append(
                f"- **{a['ticker']}** ({a['weight']*100:.0f}%): {pick_rationale}"
            )

        lines.append(f"\n**Rebalancing Plan:** {profile['rebalancing']}")

    lines.append("\n---\n")
    lines.append("## Model Comparison Summary")
    lines.append("")

    # Build comparison table
    lines.append("| Metric | " + " | ".join(portfolios.keys()) + " |")
    lines.append("|--------|" + "|".join(["--------"] * len(portfolios)) + "|")

    all_metrics = {}
    for label, port in portfolios.items():
        all_metrics[label] = compute_metrics(port, all_prices)

    metric_rows = [
        ("Strategy", lambda m, p: p.get("strategy_name", "N/A")),
        ("Current Value", lambda m, p: f"${m['current_value']:,.2f}"),
        ("Total Return", lambda m, p: f"{m['total_return_pct']:+.2%}"),
        ("Total Return ($)", lambda m, p: f"${m['total_return_dollar']:+,.2f}"),
        ("AI Generation Cost", lambda m, p: f"${m['ai_cost_usd']:.4f}" if m['ai_cost_usd'] > 0 else "Free (local)"),
        ("AI Tokens Used", lambda m, p: f"{m['ai_tokens']:,}" if m['ai_tokens'] > 0 else "N/A"),
        ("Net Return (after AI cost)", lambda m, p: f"{m['net_return_pct']:+.2%}"),
        ("Net Return $ (after AI cost)", lambda m, p: f"${m['net_return_dollar']:+,.2f}"),
        ("Annualized Return", lambda m, p: f"{m['annualized_return']:+.2%}"),
        ("Max Drawdown", lambda m, p: f"{m['max_drawdown']:.2%}"),
        ("Volatility (ann.)", lambda m, p: f"{m['volatility']:.2%}"),
        ("Sharpe Ratio", lambda m, p: f"{m['sharpe_ratio']:.2f}"),
        ("Sortino Ratio", lambda m, p: f"{m['sortino_ratio']:.2f}"),
        ("Calmar Ratio", lambda m, p: f"{m['calmar_ratio']:.2f}"),
        ("Win Rate", lambda m, p: f"{m['win_rate']:.1%}"),
        ("Profit Factor", lambda m, p: f"{m['profit_factor']:.2f}"),
        ("# Trades", lambda m, p: str(m['num_trades'])),
        ("# Holdings", lambda m, p: str(m['num_holdings'])),
        ("Cash Remaining", lambda m, p: f"${m['cash_remaining']:,.2f}"),
        ("Best Holding", lambda m, p: f"{m['best_holding']} ({m['best_holding_return']:+.2%})"),
        ("Worst Holding", lambda m, p: f"{m['worst_holding']} ({m['worst_holding_return']:+.2%})"),
    ]

    for row_name, fmt_fn in metric_rows:
        cells = []
        for label in portfolios:
            m = all_metrics[label]
            p = portfolios[label]
            cells.append(fmt_fn(m, p))
        lines.append(f"| {row_name} | " + " | ".join(cells) + " |")

    # Per-model detail sections
    for label, port in portfolios.items():
        m = all_metrics[label]
        lines.append(f"\n---\n")
        lines.append(f"## {label} — {port.get('strategy_name', 'N/A')}")
        lines.append(f"\n**Risk Level:** {port['strategy_data'].get('risk_level', 'N/A')}")
        lines.append(f"**Rationale:** {port['strategy_data'].get('rationale', 'N/A')}")
        lines.append(f"**Time Horizon:** {port['strategy_data'].get('time_horizon', 'N/A')}")

        # Holdings detail
        lines.append("\n### Current Holdings\n")
        lines.append("| Ticker | Shares | Avg Cost | Current Price | Market Value | Return |")
        lines.append("|--------|--------|----------|---------------|-------------|--------|")

        for ticker, info in port["holdings"].items():
            price = all_prices.get(ticker, 0)
            mkt_val = info["shares"] * price
            ret = (price - info["avg_cost"]) / info["avg_cost"] if info["avg_cost"] > 0 else 0
            lines.append(
                f"| {ticker} | {int(info['shares'])} | ${info['avg_cost']:.2f} | "
                f"${price:.2f} | ${mkt_val:,.2f} | {ret:+.2%} |"
            )

        # # Transaction log
        # lines.append("\n### Transaction Log\n")
        # lines.append("| Date | Action | Ticker | Shares | Price | Total | Cash After |")
        # lines.append("|------|--------|--------|--------|-------|-------|------------|")
        # for txn in port["transactions"]:
        #     lines.append(
        #         f"| {txn['timestamp'][:10]} | {txn['action']} | {txn['ticker']} | "
        #         f"{txn['shares']} | ${txn['price']:.2f} | ${txn['total_cost']:,.2f} | "
        #         f"${txn['cash_after']:,.2f} |"
        #     )

    # Footer
    lines.append("\n---\n")
    lines.append(
        "⚠️ **Disclaimer:** This tool is for learning and experimentation only. "
        "Always consult a qualified financial advisor before making real investment "
        "decisions. The strategy is automatically generated from large language models "
        "and may contain errors, omissions, or outdated information. It is provided “as is,” "
        "without warranties of any kind, express or implied. This application is not investment, "
        "legal, security, or policy advice and must not be relied upon for decision-making. "
        "You are responsible for independently verifying facts and conclusions with primary sources "
        "and qualified professionals. Neither the author nor providers of this service accept any "
        "liability for losses or harms arising from use of this content. No duty to update is assumed."
    
    )

    return "\n".join(lines)


# ── Excel report generation ────────────────────────────────────────────────────

def generate_excel_report(
    thesis: str,
    portfolios: dict,
    all_prices: dict,
    benchmark_ticker: str,
    initial_capital: float,
    start_date: str,
    filepath: str,
) -> str:
    """
    Generate a multi-sheet Excel report with full analysis.

    Sheets:
      - Summary:      Side-by-side metrics for all models
      - Per model:    Holdings + transactions
      - Raw Data:     All transactions across models

    Returns:
        The filepath written to.
    """
    wb = Workbook()

    # ── Sheet 1: Summary comparison ────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = "2F5496"

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "Investment Strategy Performance Report"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="2F5496")

    ws["A3"] = "Thesis:"
    ws["B3"] = thesis[:200] + ("..." if len(thesis) > 200 else "")
    ws["A4"] = "Initial Capital:"
    ws["B4"] = initial_capital
    ws["B4"].number_format = MONEY_FMT
    ws["A5"] = "Start Date:"
    ws["B5"] = start_date
    ws["A6"] = "Benchmark:"
    ws["B6"] = benchmark_ticker
    ws["A7"] = "Report Generated:"
    ws["B7"] = datetime.now().strftime("%Y-%m-%d %H:%M")

    # Comparison table starting at row 9
    labels = list(portfolios.keys())
    headers = ["Metric"] + labels
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=9, column=col_idx, value=h)
    _style_header(ws, 9, len(headers))

    all_metrics = {}
    for label, port in portfolios.items():
        all_metrics[label] = compute_metrics(port, all_prices)

    metric_defs = [
        ("Strategy Name", lambda m, p: p.get("strategy_name", "N/A"), None),
        ("Current Value", lambda m, p: m["current_value"], MONEY_FMT),
        ("Total Return (%)", lambda m, p: m["total_return_pct"], PCT_FMT),
        ("Total Return ($)", lambda m, p: m["total_return_dollar"], MONEY_FMT),
        ("AI Generation Cost", lambda m, p: m["ai_cost_usd"], '$#,##0.0000'),
        ("AI Tokens Used", lambda m, p: m["ai_tokens"], '#,##0'),
        ("Net Return (after AI)", lambda m, p: m["net_return_pct"], PCT_FMT),
        ("Net Return $ (after AI)", lambda m, p: m["net_return_dollar"], MONEY_FMT),
        ("Annualized Return", lambda m, p: m["annualized_return"], PCT_FMT),
        ("Max Drawdown", lambda m, p: m["max_drawdown"], PCT_FMT),
        ("Volatility (ann.)", lambda m, p: m["volatility"], PCT_FMT),
        ("Sharpe Ratio", lambda m, p: m["sharpe_ratio"], "0.00"),
        ("Sortino Ratio", lambda m, p: m["sortino_ratio"], "0.00"),
        ("Calmar Ratio", lambda m, p: m["calmar_ratio"], "0.00"),
        ("Win Rate", lambda m, p: m["win_rate"], PCT_FMT),
        ("Profit Factor", lambda m, p: m["profit_factor"], "0.00"),
        ("Days Elapsed", lambda m, p: m["days_elapsed"], None),
        ("# Trades", lambda m, p: m["num_trades"], None),
        ("# Holdings", lambda m, p: m["num_holdings"], None),
        ("Cash Remaining", lambda m, p: m["cash_remaining"], MONEY_FMT),
        ("% Invested", lambda m, p: m["invested_pct"], PCT_FMT),
        ("Best Holding", lambda m, p: m["best_holding"], None),
        ("Best Holding Return", lambda m, p: m["best_holding_return"], PCT_FMT),
        ("Worst Holding", lambda m, p: m["worst_holding"], None),
        ("Worst Holding Return", lambda m, p: m["worst_holding_return"], PCT_FMT),
    ]

    for row_offset, (metric_name, fn, fmt) in enumerate(metric_defs):
        row = 10 + row_offset
        ws.cell(row=row, column=1, value=metric_name).font = Font(name="Arial", bold=True)
        for col_offset, label in enumerate(labels):
            cell = ws.cell(row=row, column=2 + col_offset)
            cell.value = fn(all_metrics[label], portfolios[label])
            if fmt:
                cell.number_format = fmt
            cell.border = THIN_BORDER

    ws.column_dimensions["A"].width = 22
    for i in range(len(labels)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 20

    # ── Sheet: Thesis Interpretation ───────────────────────────────────────
    # Narrative summary of how each LLM converted the thesis into a strategy
    ws_interp = wb.create_sheet("Thesis Interpretation")
    ws_interp.sheet_properties.tabColor = "7030A0"

    ws_interp.merge_cells("A1:D1")
    ws_interp["A1"] = "How Each Model Interpreted the Thesis"
    ws_interp["A1"].font = Font(name="Arial", bold=True, size=14, color="7030A0")

    ws_interp["A3"] = "Original Thesis:"
    ws_interp["A3"].font = Font(name="Arial", bold=True)
    ws_interp.merge_cells("B3:D3")
    ws_interp["B3"] = thesis[:500] + ("..." if len(thesis) > 500 else "")
    ws_interp["B3"].alignment = Alignment(wrap_text=True)

    # ── Comparative narrative summary ──────────────────────────────────────
    ws_interp.merge_cells("A5:D5")
    ws_interp["A5"] = "Comparative Analysis"
    ws_interp["A5"].font = Font(name="Arial", bold=True, size=12, color="7030A0")

    # Build the narrative (strip Markdown bold markers for Excel)
    narrative = _build_comparative_narrative(portfolios).replace("**", "")
    ws_interp.merge_cells("A6:D10")
    ws_interp["A6"] = narrative
    ws_interp["A6"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_interp["A6"].font = Font(name="Arial", size=10)
    ws_interp.row_dimensions[6].height = 100  # Give the narrative cell room

    # ── Per-model translation summaries ────────────────────────────────────
    # Each model gets a short "translation" sentence + detailed breakdown
    interp_row = 12
    for label, port in portfolios.items():
        sd = port.get("strategy_data", {})
        allocs = sd.get("allocations", [])
        sorted_allocs = sorted(allocs, key=lambda a: a.get("weight", 0), reverse=True)
        profile = _build_model_profile(port)

        # ── Model header ──
        ws_interp.merge_cells(f"A{interp_row}:D{interp_row}")
        ws_interp.cell(row=interp_row, column=1,
                       value=f"{label} → {port.get('strategy_name', 'N/A')}")
        ws_interp.cell(row=interp_row, column=1).font = Font(
            name="Arial", bold=True, size=12, color="2F5496")
        interp_row += 1

        # ── Translation summary: concise narrative of how thesis → trades ──
        concentration = "concentrated" if profile["top_weight"] >= 0.30 else "diversified"
        vehicle_desc = (
            "primarily through ETFs" if profile["etf_count"] > profile["stock_count"]
            else "primarily through individual stocks"
            if profile["stock_count"] > profile["etf_count"]
            else "through a balanced mix of ETFs and individual stocks"
        )
        top_pick = profile["top_3"][0] if profile["top_3"] else None
        top_desc = (
            f", anchored by a {top_pick['weight']*100:.0f}% position in "
            f"{top_pick['ticker']} ({top_pick.get('name', top_pick['ticker'])})"
            if top_pick else ""
        )
        translation_text = (
            f"Interpreted the thesis as a {profile['risk_level'].lower()} opportunity "
            f"and expressed it {vehicle_desc} over a {profile['time_horizon'].lower()} "
            f"horizon. Built a {concentration} portfolio of {profile['num_positions']} "
            f"positions{top_desc}."
        )

        ws_interp.cell(row=interp_row, column=1, value="Translation Summary")
        ws_interp.cell(row=interp_row, column=1).font = Font(
            name="Arial", bold=True, italic=True, color="7030A0")
        ws_interp.merge_cells(f"B{interp_row}:D{interp_row}")
        ws_interp.cell(row=interp_row, column=2, value=translation_text)
        ws_interp.cell(row=interp_row, column=2).alignment = Alignment(wrap_text=True)
        ws_interp.cell(row=interp_row, column=2).font = Font(name="Arial", italic=True)
        ws_interp.row_dimensions[interp_row].height = 45
        interp_row += 1

        # ── Key attributes ──
        attrs = [
            ("Risk Posture", sd.get("risk_level", "N/A")),
            ("Time Horizon", sd.get("time_horizon", "N/A")),
            ("# Positions", str(len(allocs))),
            ("ETFs / Stocks", f"{profile['etf_count']} ETFs, {profile['stock_count']} stocks"),
            ("Rebalancing Plan", sd.get("rebalancing_notes", "N/A")),
        ]
        for attr_name, attr_val in attrs:
            ws_interp.cell(row=interp_row, column=1, value=attr_name)
            ws_interp.cell(row=interp_row, column=1).font = Font(name="Arial", bold=True)
            ws_interp.cell(row=interp_row, column=2, value=attr_val)
            interp_row += 1

        # ── Model's rationale ──
        ws_interp.cell(row=interp_row, column=1, value="Model's Rationale")
        ws_interp.cell(row=interp_row, column=1).font = Font(name="Arial", bold=True)
        ws_interp.merge_cells(f"B{interp_row}:D{interp_row}")
        ws_interp.cell(row=interp_row, column=2, value=sd.get("rationale", "N/A"))
        ws_interp.cell(row=interp_row, column=2).alignment = Alignment(wrap_text=True)
        interp_row += 1

        # ── Position-level reasoning table ──
        interp_row += 1
        pos_headers = ["Ticker", "Weight", "Name", "Reasoning"]
        for col_idx, h in enumerate(pos_headers, 1):
            ws_interp.cell(row=interp_row, column=col_idx, value=h)
        _style_header(ws_interp, interp_row, len(pos_headers))
        interp_row += 1

        for a in sorted_allocs:
            ws_interp.cell(row=interp_row, column=1, value=a["ticker"])
            ws_interp.cell(row=interp_row, column=2, value=a["weight"])
            ws_interp.cell(row=interp_row, column=2).number_format = PCT_FMT
            ws_interp.cell(row=interp_row, column=3, value=a.get("name", a["ticker"]))
            ws_interp.cell(row=interp_row, column=4,
                           value=a.get("rationale", "No specific reason given."))
            ws_interp.cell(row=interp_row, column=4).alignment = Alignment(wrap_text=True)
            for c in range(1, 5):
                ws_interp.cell(row=interp_row, column=c).border = THIN_BORDER
            interp_row += 1

        interp_row += 2  # Spacing between models

    ws_interp.column_dimensions["A"].width = 20
    ws_interp.column_dimensions["B"].width = 30
    ws_interp.column_dimensions["C"].width = 25
    ws_interp.column_dimensions["D"].width = 55

    # ── Per-model sheets ───────────────────────────────────────────────────
    # openpyxl forbids these characters in sheet titles: \ / * ? : [ ]
    import re as _re
    _INVALID_SHEET_CHARS = _re.compile(r'[\\/*?:\[\]]')

    used_sheet_names = set()
    for label, port in portfolios.items():
        # Strip forbidden characters, then truncate to 28 chars (Excel max is 31)
        safe_name = _INVALID_SHEET_CHARS.sub("", label)[:28]
        # Deduplicate: append a suffix if the truncated name already exists
        if safe_name in used_sheet_names:
            suffix = 2
            while f"{safe_name[:26]}_{suffix}" in used_sheet_names:
                suffix += 1
            safe_name = f"{safe_name[:26]}_{suffix}"
        used_sheet_names.add(safe_name)
        ws_model = wb.create_sheet(safe_name)

        # Holdings section
        ws_model["A1"] = f"{label} — Holdings"
        ws_model["A1"].font = Font(name="Arial", bold=True, size=12, color="2F5496")

        h_headers = ["Ticker", "Shares", "Avg Cost", "Current Price",
                      "Market Value", "Gain/Loss ($)", "Return (%)"]
        for col_idx, h in enumerate(h_headers, 1):
            ws_model.cell(row=3, column=col_idx, value=h)
        _style_header(ws_model, 3, len(h_headers))

        row = 4
        for ticker, info in port["holdings"].items():
            price = all_prices.get(ticker, 0)
            mkt_val = info["shares"] * price
            gl = mkt_val - (info["shares"] * info["avg_cost"])
            ret = (price - info["avg_cost"]) / info["avg_cost"] if info["avg_cost"] > 0 else 0

            ws_model.cell(row=row, column=1, value=ticker)
            ws_model.cell(row=row, column=2, value=int(info["shares"]))
            ws_model.cell(row=row, column=3, value=info["avg_cost"]).number_format = MONEY_FMT
            ws_model.cell(row=row, column=4, value=price).number_format = MONEY_FMT
            ws_model.cell(row=row, column=5, value=mkt_val).number_format = MONEY_FMT
            ws_model.cell(row=row, column=6, value=gl).number_format = MONEY_FMT
            ws_model.cell(row=row, column=7, value=ret).number_format = PCT_FMT
            row += 1

        # Cash row
        ws_model.cell(row=row, column=1, value="CASH").font = Font(bold=True)
        ws_model.cell(row=row, column=5, value=port["cash"]).number_format = MONEY_FMT

        # Transaction section
        txn_start = row + 3
        ws_model.cell(row=txn_start - 1, column=1, value=f"{label} — Transactions")
        ws_model.cell(row=txn_start - 1, column=1).font = Font(name="Arial", bold=True, size=12, color="2F5496")

        t_headers = ["Date", "Action", "Ticker", "Shares", "Price",
                      "Total", "Cash After", "Notes"]
        for col_idx, h in enumerate(t_headers, 1):
            ws_model.cell(row=txn_start, column=col_idx, value=h)
        _style_header(ws_model, txn_start, len(t_headers))

        for i, txn in enumerate(port["transactions"]):
            r = txn_start + 1 + i
            ws_model.cell(row=r, column=1, value=txn["timestamp"][:10])
            ws_model.cell(row=r, column=2, value=txn["action"])
            ws_model.cell(row=r, column=3, value=txn["ticker"])
            ws_model.cell(row=r, column=4, value=txn["shares"])
            ws_model.cell(row=r, column=5, value=txn["price"]).number_format = MONEY_FMT
            ws_model.cell(row=r, column=6, value=txn["total_cost"]).number_format = MONEY_FMT
            ws_model.cell(row=r, column=7, value=txn["cash_after"]).number_format = MONEY_FMT
            ws_model.cell(row=r, column=8, value=txn.get("notes", ""))

        for col, w in zip("ABCDEFGH", [10, 8, 10, 8, 14, 14, 14, 30]):
            ws_model.column_dimensions[col].width = w

    # ── All Transactions sheet ─────────────────────────────────────────────
    ws_all = wb.create_sheet("All Transactions")
    all_headers = ["Model", "Date", "Action", "Ticker", "Shares",
                   "Price", "Total", "Cash After", "Notes"]
    for col_idx, h in enumerate(all_headers, 1):
        ws_all.cell(row=1, column=col_idx, value=h)
    _style_header(ws_all, 1, len(all_headers))

    row = 2
    for label, port in portfolios.items():
        for txn in port["transactions"]:
            ws_all.cell(row=row, column=1, value=label)
            ws_all.cell(row=row, column=2, value=txn["timestamp"][:10])
            ws_all.cell(row=row, column=3, value=txn["action"])
            ws_all.cell(row=row, column=4, value=txn["ticker"])
            ws_all.cell(row=row, column=5, value=txn["shares"])
            ws_all.cell(row=row, column=6, value=txn["price"]).number_format = MONEY_FMT
            ws_all.cell(row=row, column=7, value=txn["total_cost"]).number_format = MONEY_FMT
            ws_all.cell(row=row, column=8, value=txn["cash_after"]).number_format = MONEY_FMT
            ws_all.cell(row=row, column=9, value=txn.get("notes", ""))
            row += 1

    for col, w in zip("ABCDEFGHI", [18, 12, 8, 10, 8, 14, 14, 14, 30]):
        ws_all.column_dimensions[col].width = w

    wb.save(filepath)
    return filepath
