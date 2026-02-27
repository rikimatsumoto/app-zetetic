"""
portfolio_manager.py
====================
Handles creation and export of the Excel portfolio file.
Internal state tracking is handled by transaction_store.py;
this module creates the downloadable Excel representation.
"""

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── Styling ────────────────────────────────────────────────────────────────────
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill("solid", fgColor="2F5496")
MONEY_FMT = '$#,##0.00'
PCT_FMT = '0.00%'
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_header(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def export_portfolio_to_excel(portfolio: dict, prices: dict, filepath: str) -> str:
    """
    Export a portfolio state dict (from transaction_store) to a formatted Excel file.

    Creates three sheets:
      - Portfolio Summary: current holdings with live prices
      - Transactions: full trade log
      - Performance: snapshot history

    Args:
        portfolio: Portfolio state dict from transaction_store.
        prices:    {ticker: current_price} mapping.
        filepath:  Where to write the .xlsx file.

    Returns:
        The filepath written to.
    """
    wb = Workbook()

    # ── Sheet 1: Portfolio Summary ─────────────────────────────────────────
    ws = wb.active
    ws.title = "Portfolio Summary"
    ws.sheet_properties.tabColor = "2F5496"

    ws.merge_cells("A1:G1")
    ws["A1"] = f"Investment Portfolio — {portfolio.get('strategy_name', 'Strategy')}"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="2F5496")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A3"] = "Model:"
    ws["B3"] = portfolio.get("model_label", "N/A")
    ws["A4"] = "Initial Capital:"
    ws["B4"] = portfolio["initial_capital"]
    ws["B4"].number_format = MONEY_FMT
    ws["A5"] = "Start Date:"
    ws["B5"] = portfolio.get("start_date", "N/A")
    ws["A6"] = "Generated:"
    ws["B6"] = datetime.now().strftime("%Y-%m-%d %H:%M")

    headers = ["Ticker", "Shares", "Avg Cost", "Current Price",
               "Market Value", "Gain/Loss ($)", "Gain/Loss (%)"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=8, column=col_idx, value=h)
    _style_header(ws, 8, len(headers))

    row = 9
    for ticker, info in portfolio["holdings"].items():
        price = prices.get(ticker.upper(), 0)
        shares = info["shares"]
        avg_cost = info["avg_cost"]
        mkt_val = round(shares * price, 2)
        gain_dollar = round(mkt_val - (shares * avg_cost), 2)
        gain_pct = (price - avg_cost) / avg_cost if avg_cost > 0 else 0

        ws.cell(row=row, column=1, value=ticker)
        ws.cell(row=row, column=2, value=int(shares))
        ws.cell(row=row, column=3, value=avg_cost).number_format = MONEY_FMT
        ws.cell(row=row, column=4, value=price).number_format = MONEY_FMT
        ws.cell(row=row, column=5, value=mkt_val).number_format = MONEY_FMT
        ws.cell(row=row, column=6, value=gain_dollar).number_format = MONEY_FMT
        ws.cell(row=row, column=7, value=gain_pct).number_format = PCT_FMT
        for c in range(1, 8):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    # Cash row
    ws.cell(row=row, column=1, value="CASH").font = Font(name="Arial", bold=True)
    ws.cell(row=row, column=5, value=portfolio["cash"]).number_format = MONEY_FMT

    # Total row
    total_row = row + 2
    ws.cell(row=total_row, column=4, value="TOTAL:").font = Font(name="Arial", bold=True, size=12)
    total_val = portfolio["cash"]
    for ticker, info in portfolio["holdings"].items():
        p = prices.get(ticker.upper(), 0)
        total_val += info["shares"] * p
    ws.cell(row=total_row, column=5, value=round(total_val, 2))
    ws.cell(row=total_row, column=5).number_format = MONEY_FMT
    ws.cell(row=total_row, column=5).font = Font(name="Arial", bold=True, size=12)

    for col, w in zip("ABCDEFG", [12, 10, 14, 14, 16, 16, 14]):
        ws.column_dimensions[col].width = w

    # ── Sheet 2: Transactions ──────────────────────────────────────────────
    ws_txn = wb.create_sheet("Transactions")
    ws_txn.sheet_properties.tabColor = "548235"

    txn_headers = ["Date", "Action", "Ticker", "Shares", "Price",
                   "Total Cost", "Cash Before", "Cash After", "Notes"]
    for col_idx, h in enumerate(txn_headers, 1):
        ws_txn.cell(row=1, column=col_idx, value=h)
    _style_header(ws_txn, 1, len(txn_headers))

    for i, txn in enumerate(portfolio["transactions"]):
        r = 2 + i
        ws_txn.cell(row=r, column=1, value=txn["timestamp"])
        ws_txn.cell(row=r, column=2, value=txn["action"])
        ws_txn.cell(row=r, column=3, value=txn["ticker"])
        ws_txn.cell(row=r, column=4, value=txn["shares"])
        ws_txn.cell(row=r, column=5, value=txn["price"]).number_format = MONEY_FMT
        ws_txn.cell(row=r, column=6, value=txn["total_cost"]).number_format = MONEY_FMT
        ws_txn.cell(row=r, column=7, value=txn["cash_before"]).number_format = MONEY_FMT
        ws_txn.cell(row=r, column=8, value=txn["cash_after"]).number_format = MONEY_FMT
        ws_txn.cell(row=r, column=9, value=txn.get("notes", ""))

    for col, w in zip("ABCDEFGHI", [18, 8, 10, 8, 14, 14, 14, 14, 30]):
        ws_txn.column_dimensions[col].width = w

    # ── Sheet 3: Performance Snapshots ─────────────────────────────────────
    ws_perf = wb.create_sheet("Performance")
    ws_perf.sheet_properties.tabColor = "BF8F00"

    perf_headers = ["Date", "Portfolio Value", "Benchmark Value"]
    for col_idx, h in enumerate(perf_headers, 1):
        ws_perf.cell(row=1, column=col_idx, value=h)
    _style_header(ws_perf, 1, len(perf_headers))

    for i, snap in enumerate(portfolio.get("performance_snapshots", [])):
        r = 2 + i
        ws_perf.cell(row=r, column=1, value=snap["date"])
        ws_perf.cell(row=r, column=2, value=snap["portfolio_value"]).number_format = MONEY_FMT
        ws_perf.cell(row=r, column=3, value=snap["benchmark_value"]).number_format = MONEY_FMT

    for col, w in zip("ABC", [14, 18, 18]):
        ws_perf.column_dimensions[col].width = w

    wb.save(filepath)
    return filepath
